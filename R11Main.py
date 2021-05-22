import sys
import os.path
from openpyxl import load_workbook
import datetime
from R11AusenciaAbonada import AusenciaAbonada
from R11ConsultaMedica import ConsultaMedica
from R11RegrasGerais import RegrasGerais
from R11RegrasReprovacao import RegrasReprovacao
from R11 import CONST



def getColumnIndexes(sheet, colunas):
    maxcolunms = sheet.max_column + 1
    indexesColunas = {}
    for c in range(1, maxcolunms):
        cell = sheet.cell(row=1, column=c)
        if cell.value in colunas:
            indexesColunas[cell.value] = c
    return indexesColunas


def main(args):
    if len(args)==0:
        return
    if not os.path.isfile(args[0]):
        return

    workBook = load_workbook(args[0])
    planilhaAusenciaAbonada = workBook[CONST.NOME_AUSENCIA_ABONADA]
    planilhaConsultaMedica = workBook[CONST.NOME_CONSULTA_MEDICA]
    indicesAusenciaAbonada = getColumnIndexes( planilhaAusenciaAbonada, CONST.COLUNAS_AUSENCIA_ABONADA)
    indicesConsultaMedica = getColumnIndexes( planilhaConsultaMedica, CONST.COLUNAS_CONSULTA_MEDICA)

    #ausenciaAbonada = AusenciaAbonada(planilhaAusenciaAbonada, indicesAusenciaAbonada)
    #ausenciaAbonada.validar()
    consultaMedica = ConsultaMedica(planilhaConsultaMedica, indicesConsultaMedica)
    consultaMedica.validar()

    workBook.save(args[0])
    k = input("Pressione ENTER para encerrar...")

if __name__ == '__main__':
    #main(sys.argv[1:])
    main([r'c:\RPA\Planilhas\Planilha de Base.xlsx'])
