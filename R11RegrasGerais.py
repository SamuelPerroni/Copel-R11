from io import SEEK_CUR

from openpyxl.descriptors import serialisable
from R11 import CONST
from datetime import datetime, timedelta

class RegrasGerais:
    def __init__(self, planilha, indices):
        self.indices = indices
        self.planilha = planilha

    def is_integer(self, n):
        try:
            float(n)
        except ValueError:
            return False
        else:
            return float(n).is_integer()

    def stringToTime(self, hora):
        return datetime.strptime(hora,'%H:%M')

    def diffHours(self, horaMaior, horaMenor):
        """EX.: horaMaior= "21:00" - horaMenor= "16:30" """
        return (self.stringToTime(horaMaior) - self.stringToTime(horaMenor)).seconds/60/60

    def horasTuplas(self, marcacoes):
        return [tuple(x.split(" - ")) for x in marcacoes.split(";")]

    def somaHorasMarcacoes(self, marcacoes):
        """horas (string de horas. Ex.: "09:00 - 13:00;14:00 - 16:00")"""
        total = 0.0
        if not marcacoes == CONST.SEM_DADOS:
            range_marcacoes = self.horasTuplas(marcacoes)
            for marcacao in range_marcacoes:
                total += self.diffHours(marcacao[1], marcacao[0])
        return total

    def totalHorasPorPeriodo(self, marcacoes, ausencia, presenca, intervaloPeriodo):
        total = 0.0
        range = self.horasTuplas(marcacoes)
        if not ausencia == CONST.SEM_DADOS:
            range += self.horasTuplas(ausencia)
        if not presenca == CONST.SEM_DADOS:
            range += self.horasTuplas(presenca)
        limite1 = intervaloPeriodo[0:intervaloPeriodo.find(" -")]
        limite2 = intervaloPeriodo[intervaloPeriodo.find(" -")+3:]
        range1 = ';'.join(("{} - {}".format(x[0],x[1]) for x in range if x[0] <= limite1))
        range2 = ';'.join(("{} - {}".format(x[0],x[1]) for x in range if limite2 <= x[1]))
        retorno = [0.0, 0.0]
        retorno[0] = self.somaHorasMarcacoes(range1) if not range1 == '' else 0.0
        retorno[1] = self.somaHorasMarcacoes(range2) if not range2 == '' else 0.0

        return retorno


    def horasTotal(self, marcacoes, ausencia, presenca):
        return self.somaHorasMarcacoes(marcacoes) + self.somaHorasMarcacoes(ausencia) + self.somaHorasMarcacoes(presenca)

    def getVal(self, linha, indice):
        return self.planilha.cell(row=linha, column=self.indices[indice]).value

    def reprovar(self, linha, valor):
        self.planilha.cell(row=linha, column=self.indices["RESULTADO"]).value = "REPROVADO"
        self.planilha.cell(row=linha, column=self.indices["SOLUCAO_RESPOSTA"]).value = valor
    
    def abonar(self, linha, valor):
        self.planilha.cell(row=linha, column=self.indices["RESULTADO"]).value = CONST.ABONADO
        self.planilha.cell(row=linha, column=self.indices["SOLUCAO_RESPOSTA"]).value = valor

    def reprovado(self, linha):
        return self.getVal(linha,"RESULTADO") == CONST.REPROVADO

    def semDados(self, linha):
        return (self.getVal(linha,"AUSENCIA") == CONST.SEM_DADOS 
                and self.getVal(linha,"PRESENCA") == CONST.SEM_DADOS 
                and self.getVal(linha,"QTD_MARCACOES") == CONST.SEM_DADOS)
    
    def __marcacoesPorPeriodo(self, marcacoes, periodo):
        per = periodo.split(" - ")
        range_marcacoes = self.horasTuplas(marcacoes)
        retorno_marcacoes = ""
        sep = ""
        for marcacao in range_marcacoes:
            if  per[0] <= marcacao[0] <= per[1] or per[0] <= marcacao[1] <= per[1]:        
                retorno_marcacoes += sep + marcacao[0] + " - " + marcacao[1]
                sep = ";"
        return retorno_marcacoes

    def __horasEntrePeriodos(self, hora, periodos):
        """Retorna True or False se uma determinada hora está entre dois horários.
           Pametros: hora (string, ex: 23:30) - periodos (string de horas, ex: "09:00 - 13:00;14:00 - 16:00")"""
        range_periodos =self.horasTuplas(periodos)
        for periodo in range_periodos:
            if  periodo[0] <= hora <= periodo[1]:
                return True
        return False
    

    def __horasEntrePeriodosExcetoInicioFim(self, hora, periodos):
        """Retorna True or False se uma determinada hora está entre dois horários.
           Pametros: hora (string, ex: 23:30) - periodos (string de horas, ex: "09:00 - 13:00;14:00 - 16:00")"""
        range_periodos = self.horasTuplas(periodos)
        for periodo in range_periodos:
            if  periodo[0] < hora < periodo[1]:
                return True
        return False
    
    def __marcacoesEntrePeriodos(self, marcacoes, periodos):
        """Retorna True or False se uma determinada hora está entre dois horários.
           Pametros: marcacoes (string, ex: "09:00 - 13:00;14:00 - 16:00") - periodos (string de horas, ex: "09:00 - 13:00;14:00 - 16:00")"""
        range_periodos = self.horasTuplas(periodos)
        range_marcacoes = self.horasTuplas(marcacoes)
        for periodo in range_periodos:
            for marcacao in range_marcacoes:
                if  periodo[0] <= marcacao[0] < marcacao[1] <= periodo[1]:
                    return True
        return False

    def __retornaMarcacoesEntrePeriodos(self, marcacoes, periodos):
        """Retorna True or False se uma determinada hora está entre dois horários.
           Pametros: marcacoes (string, ex: "09:00 - 13:00;14:00 - 16:00") - periodos (string de horas, ex: "09:00 - 13:00;14:00 - 16:00")"""
        range_periodos = self.horasTuplas(periodos)
        range_marcacoes = self.horasTuplas(marcacoes)
        marcacoes = []
        for periodo in range_periodos:
            marcacoes.append([marcacao for marcacao in range_marcacoes if  periodo[0] <= marcacao[0] < marcacao[1] <= periodo[1]])
        return marcacoes

    def horaInicialForaDoPeriodo(self, linha):
        horaInicial = self.getVal(linha,"HORA_INICIAL")
        tempoTeorico = self.getVal(linha,"TEMPO_TEORICO")        
        return not self.__horasEntrePeriodos(horaInicial, tempoTeorico)

    def marcacaoImpar(self, linha):
        return (self.getVal(linha, "QTD_MARCACOES") == CONST.SEM_DADOS 
            and not int(self.getVal(linha, "QTD_MARCACOES")) % 2 == 0)

    def jornadaCumpridaDentroDoTempoTeorico(self, linha):
        # Aqui considerar as marcações de AUSENCIA E PRESENCA
        ausencia = self.getVal(linha,"AUSENCIA")
        presenca = self.getVal(linha,"PRESENCA")
        horasTeorico = self.getVal(linha,"HORAS_TEORICO").replace(',','.')  
        tempoTeorico =  self.getVal(linha,"TEMPO_TEORICO")
        marcacoes = self.getVal(linha,"MARCACOES")
        horasTotal = self.horasTotal(marcacoes, ausencia, presenca)
        dentroDoTempoTeorico = self.__marcacoesEntrePeriodos(marcacoes, tempoTeorico)
        return horasTotal >= float(horasTeorico) and dentroDoTempoTeorico
    
    def semRegistrosNaEscalaDeIntervalo(self, linha):
        """
        Se HORAS_TEORICO != 4,00
        Utilizar o valores entre os periodos: 08:30 - 11:30;14:00 - 17:00 => 11:30 - 14:00        
        """
        if self.getVal(linha,"HORAS_TEORICO") == "4,00":
            return False
        
        periodo1 = self.getVal(linha,"PERIODO1")
        periodo2 = self.getVal(linha,"PERIODO2")
        if (not periodo1 is None or not periodo1 == "") and (not periodo1 is None or not periodo1 == ""):
            periodoIntervalo = periodo1.split(" - ")[1] + " - " + periodo2.split(" - ")[0]
            marcacoes = self.getVal(linha,"MARCACOES")
            return self.__marcacoesEntrePeriodos(marcacoes, periodoIntervalo)

        return True
    
    def recuperarInicioFimIntervalo(self, periodo1, periodo2):
        if (not periodo1 is None or not periodo1 == "") and (not periodo1 is None or not periodo1 == ""):
            return periodo1.split(" - ")[1] + " - " + periodo2.split(" - ")[0]
        return None

    def validarHoraInicialIntervalo(self, linha):
        horaInicial = self.getVal(linha,"HORA_INICIAL")
        tempoTeorico = self.getVal(linha,"TEMPO_TEORICO")        
        return self.__horasEntrePeriodos(horaInicial, tempoTeorico)
 
    def retornaGaps4Horas(self, marcacoes, tempoTeorico):
        m = self.horasTuplas(marcacoes)
        tt = self.horasTuplas(tempoTeorico)[0]
        gaps = []
        if tt[0] < m[0][0]:
            gaps.append("{} - {}".format(tt[0],m[0][0]))
        if len(m)>2:
            for x in range(0, len(m)-2):
                if  m[x][1] < m[x + 1][0]:          
                    gaps.append("{} - {}".format(m[x][1],m[x + 1][0]))
        if len(m)==2:
            if  m[0][1] < m[1][0]:          
                gaps.append("{} - {}".format(m[0][1],m[1][0]))

        if m[len(m)-1][1] < tt[1]:
            gaps.append("{} - {}".format(m[len(m)-1][1], tt[1]))

        return gaps


    def retornarMaiorGapPorIntervalos(self, marcacoes, tempoTeorico, intervaloPeriodo):
        ip = intervaloPeriodo.split(" - ")
        tt = self.horasTuplas(tempoTeorico)[0]
        gaps = []

        per1 = [x for x in self.horasTuplas(marcacoes) if x[0] < ip[0] <= x[1]]
        per2 = [x for x in self.horasTuplas(marcacoes) if x[0] <= ip[1] < x[1]]

        if len(per1) == 0:
            gaps.append("{} - {}".format(tt[0],ip[0]))
        else:
            per1_Inicio = per1[0][0]
            per1_Fim =  ip[0] if per1[len(per1)-1][1]< ip[0] else per1[len(per1)-1][1]
            df = 0
            if ip[0] > per1_Fim:
                df = self.diffHours(ip[0], per1_Fim)
            if self.diffHours(per1_Inicio, tt[0]) > df:
                gaps.append("{} - {}".format(tt[0], per1_Inicio))
            else:
                gaps.append("{} - {}".format(per1_Fim, ip[0]))
        if len(per2) == 0:
            gaps.append("{} - {}".format(ip[1],tt[1]))
        else:
            per2_Inicio = ip[1] if per2[0][0] > ip[1] else per2[0][0]
            per2_Fim =  per2[len(per2)-1][1]
            df1 = 0
            if ip[1] > per2_Inicio:
                df1 = self.diffHours(ip[1], per2_Inicio)
            if self.diffHours(tt[1], per2_Fim) > df1:
                gaps.append("{} - {}".format(per2_Fim, tt[1]))
            else:
                gaps.append("{} - {}".format(ip[0], per2_Inicio))
        
        return gaps

    def calcularHorasAbono(self, periodoDisponivel, horasDecimalAbonar, depois):
        hora = int(horasDecimalAbonar)
        min = (horasDecimalAbonar*60) % 60
        delta = timedelta(hours=hora, minutes=min)
        per = periodoDisponivel.split(" - ")
        marcacao = ""
        if depois:
            novaHora = self.stringToTime(per[0]) + delta
            marcacao = "{} - {}".format(per[0], datetime.strftime(novaHora, "%H:%M"))
        else:
            novaHora = self.stringToTime(per[1]) - delta
            marcacao = "{} - {}".format(datetime.strftime(novaHora, "%H:%M"), per[1])
        return marcacao
        
    def periodoDaSolicitacao(self, horaInicial, intervaloPeriodo):
        ip = intervaloPeriodo.split(" - ")
        return 0 if ip[0] < horaInicial < ip[1] else (1 if horaInicial <= ip[0] else 2)

    def calcularAbono(self, linha):
        horaInicial = self.getVal(linha,"HORA_INICIAL")
        ausencia = self.getVal(linha,"AUSENCIA")
        presenca = self.getVal(linha,"PRESENCA")
        #Hora solicitada entre AUSENCIAS
        if self.__horasEntrePeriodos(horaInicial, ausencia) and not ausencia == CONST.SEM_DADOS:
            return [CONST.REPROVADO,"12"]

        #hora solicitada entre PRESENCAS
        if not presenca == CONST.SEM_DADOS and self.__horasEntrePeriodosExcetoInicioFim(horaInicial, presenca):
            horaFinal = (datetime.strptime(horaInicial, "%H:%M") + timedelta(minutes=1)).strftime("%H:%M")
            return [CONST.ABONADO, "{} - {}".format(horaInicial,horaFinal)]
            
        marcacoes = self.getVal(linha,"MARCACOES")
        horasTeorico = self.getVal(linha,"HORAS_TEORICO")   
        tempoTeorico = self.getVal(linha,"TEMPO_TEORICO")  
        periodo1 = self.getVal(linha,"PERIODO1")
        periodo2 = self.getVal(linha,"PERIODO2")
        tt = tempoTeorico.split(" - ")
        p1 = periodo1.split(" - ")
        p2 = periodo2.split(" - ")
        hrTeorico = float(horasTeorico.replace(',','.'))
        horasTotal = self.horasTotal(marcacoes, ausencia, presenca)
        if horasTeorico == "4,00":
            if horasTotal < hrTeorico:
                gaps = self.retornaGaps4Horas(marcacoes, tempoTeorico)
                if len(gaps)>0:
                    abono = ";".join(gaps)
                    return [CONST.ABONADO, abono]
        else:                        
            if horasTotal < hrTeorico:
                intervaloPeriodo = self.recuperarInicioFimIntervalo(periodo1, periodo2)            
                periodoDaSolicitacao = self.periodoDaSolicitacao(horaInicial, intervaloPeriodo)
                totalHorasRealizadas = self.totalHorasPorPeriodo(marcacoes, ausencia, presenca, intervaloPeriodo)
                gaps = self.retornarMaiorGapPorIntervalos(marcacoes, tempoTeorico, intervaloPeriodo)

                abono = ""
                if periodoDaSolicitacao == 0:
                    if totalHorasRealizadas[0] < hrTeorico/2 and totalHorasRealizadas[1] < hrTeorico/2:
                        abono = self.calcularHorasAbono(gaps[0], (hrTeorico - horasTotal)/2, False)
                        abono = self.calcularHorasAbono(gaps[1], (hrTeorico - horasTotal)/2, True)
                    elif totalHorasRealizadas[0] < hrTeorico/2: 
                        abono = self.calcularHorasAbono(gaps[0], (hrTeorico - horasTotal)/2, False)
                    else:
                        abono = self.calcularHorasAbono(gaps[1], (hrTeorico - horasTotal)/2, True)
                elif periodoDaSolicitacao == 1:
                    abono = self.calcularHorasAbono(gaps[0], hrTeorico - horasTotal, False)
                else:
                    abono = self.calcularHorasAbono(gaps[1], hrTeorico - horasTotal, True )
                
                return [CONST.ABONADO, abono]
           

        return [CONST.REPROVADO,"13"]